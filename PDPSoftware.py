# Libraries 
import pandas as pd
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
import matplotlib.pyplot as plt
from matplotlib.ticker import StrMethodFormatter

# Progam 1 Functions: Data Processing

# Convert 392800 number to print_key format (xxx.xx-xx-xxx.x)
def convert_number(code: str) -> str:
    if not code:
        return None
    clean_code = re.sub(r"[.\-/]", "-", str(code))
    parts = clean_code.split("-")
    numbers = [p for p in parts if p.isdigit()]
    if len(numbers) < 7:
        return None
    part1 = str(int(numbers[1]))
    part2 = f"{int(numbers[2]):02d}"
    part3 = str(int(numbers[3]))
    part4 = str(int(numbers[4]))
    extension = int(numbers[5])
    if extension != 0:
        part4 = f"{part4}.{extension}"
    return f"{part1}.{part2}-{part3}-{part4}"

# Ensures no data or formatting is overwritten
def dfnooverwrite(df, filename):
    wb = load_workbook(filename)
    ws = wb.active
    for c_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=c_idx).value = col_name
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value
    wb.save(filename)

# Apply formatting and formulas to the Excel file
def apply_formatting(filename, inserted_indexes):
    wb = load_workbook(filename)
    ws = wb.active
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    exclude_headers = ["Difference (Sales Price)", "Difference (AV)"]
    for idx in inserted_indexes:
        col_num = idx + 1
        header_value = ws.cell(row=1, column=col_num).value
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = header_align
        if header_value not in exclude_headers:
            cell.fill = header_fill
    # Difference Formulas
    for row in range(2, ws.max_row + 1):
        ws[f"L{row}"] = f"=K{row}-J{row}"
        ws[f"O{row}"] = f"=N{row}-M{row}"
    # Format issue with Living Sqft column
    living_sqft_col = None
    for i, c in enumerate(ws[1], start=1):
        if c.value == "Living Sqft (Est)":
            living_sqft_col = i
            break
    if living_sqft_col:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=living_sqft_col).number_format = "General"
    # Currency Formatting
    currency_cols = ["K", "L", "M", "N", "O", "P"]
    currency_fmt = '$#,##0;[Red]($#,##0)'
    for col in currency_cols:
        for r in range(2, ws.max_row + 1):
            ws[f"{col}{r}"].number_format = currency_fmt
    # Highlight Missing Data
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    null_headers = ["5217 Sales Price", "5217 Assessed Value", "Current Assessed Value", "Condition Code"]
    null_cols = [idx + 1 for idx, cell in enumerate(ws[1]) if cell.value in null_headers]
    for col in null_cols:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value in (None, "", "nan", "Nan"):
                cell.fill = highlight_fill
    address_col = None
    verified_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Address":
            address_col = idx
        elif cell.value == "Verified (Y/N)":
            verified_col = idx
    # Highlight addresses with "N" in Verified column
    if address_col and verified_col:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=verified_col).value == "N":
                ws.cell(row=row, column=address_col).fill = highlight_fill
    wb.save(filename)
# Main Data Processing Function
def run_processing():
    # Select Files
    main_excel = filedialog.askopenfilename(title="Select Residential Sales Excel File to Edit", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not main_excel:
        return
    roll_file = filedialog.askopenfilename(title="Select Roll Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not roll_file:
        return
    sales_file = filedialog.askopenfilename(title="Select Sales Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not sales_file:
        return
    # Read Files
    df = pd.read_excel(main_excel, dtype=str)
    roll_excel = pd.read_excel(roll_file, dtype=str)
    sales_excel = pd.read_excel(sales_file, dtype=str)
    # Insert New Columns
    df.insert(10, "5217 Sales Price", "")
    df.insert(11, "Difference (Sales Price)", "")
    df.insert(12, "5217 Assessed Value", "")
    df.insert(13, "Current Assessed Value", "")
    df.insert(14, "Difference (AV)", "")
    df.insert(23, "Verified (Y/N)", "")
    df.insert(24, "Condition Code", "")
    inserted_columns = [10, 11, 12, 13, 14, 23, 24]
    for dfx, col in [(sales_excel, "print_key"), (roll_excel, "print_key")]:
        dfx[col] = dfx[col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    # Create Lookup Dictionaries
    deedprice_lookup = sales_excel.set_index("print_key")["sale_price"].to_dict()
    deedav_lookup = sales_excel.set_index("print_key")["total_av"].to_dict()
    setav_lookup = roll_excel.set_index("print_key")["total_av"].to_dict()
    cond_lookup = sales_excel.set_index("print_key")["sale_condition_code"].to_dict()
    # Data written into New Columns
    df["5217 Sales Price"] = df["Parcel Number"].apply(lambda x: deedprice_lookup.get(convert_number(x)))
    df["5217 Assessed Value"] = df["Parcel Number"].apply(lambda x: deedav_lookup.get(convert_number(x)))
    df["Current Assessed Value"] = df["Parcel Number"].apply(lambda x: setav_lookup.get(convert_number(x)))
    df["Condition Code"] = df["Parcel Number"].apply(lambda x: cond_lookup.get(convert_number(x)))
    # Data Cleaning and Verification
    error_input_cols = [
        "5217 Sales Price","5217 Assessed Value","Current Assessed Value","Current Price","Year Built",
        "Lot Size Acres","Living Sqft (Est)","Tax Assessed Value","Bedrooms Total","Bathrooms Full",
        "Bathrooms Half","Association Fee"
    ]
    for col in error_input_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    if "MLS #" in df.columns:
        df["MLS #"] = df["MLS #"].apply(lambda x: int(x) if isinstance(x, str) and x.isdigit() else x)
    # Verified Column
    def check_verified(row):
        return "Y" if (
            pd.notna(row["5217 Sales Price"]) and
            pd.notna(row["5217 Assessed Value"]) and
            pd.notna(row["Current Assessed Value"]) and
            pd.notna(row["Condition Code"])
        ) else "N"
    df["Verified (Y/N)"] = df.apply(check_verified, axis=1)
    dfnooverwrite(df, main_excel)
    apply_formatting(main_excel, inserted_columns)
    messagebox.showinfo("Finished", "Excel Processing Complete!")

# Program 2 Functions: AnnualSales Tools

# Remove "Ghost Rows" to prevent data being appended incorrectly
def remove_ghost_rows(ws, data_columns=25):
    max_row = ws.max_row
    while max_row > 1:
        row = ws[max_row][:data_columns]
        if any(cell.value not in (None, "") for cell in row):
            break
        ws.delete_rows(max_row)
        max_row -= 1
# Append Rows from Source to Destination
def append_rows():
    temp_root = tk.Tk()
    temp_root.withdraw()
    source_file = filedialog.askopenfilename(parent=temp_root, title="Select Source Excel File", filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls")])
    if not source_file:
        temp_root.destroy()
        return
    try:
        wb_source = openpyxl.load_workbook(source_file)
        ws_source = wb_source.active
    except Exception as e:
        temp_root.destroy()
        messagebox.showerror("Error", f"Unable to open source file.\n\n{e}")
        return
    dest_file = "Annual Sales.xlsx"
    try:
        wb_dest = openpyxl.load_workbook(dest_file)
    except Exception as e:
        temp_root.destroy()
        messagebox.showerror("Error", f"{dest_file} not found or cannot be opened.\n\n{e}")
        return
    sheet_names = wb_dest.sheetnames
    sheet_list_text = "Available sheets:\n" + "\n".join(sheet_names) + "\n\nEnter destination sheet name:"
    sheet_selection = simpledialog.askstring("Select Sheet", sheet_list_text, parent=temp_root)
    temp_root.destroy()
    if not sheet_selection or sheet_selection not in wb_dest.sheetnames:
        messagebox.showerror("Error", f"'{sheet_selection}' does not exist in {dest_file}")
        return
    ws_dest = wb_dest[sheet_selection]
    remove_ghost_rows(ws_dest, data_columns=25)
    start_row = ws_dest.max_row + 1
    # Adding New Columns with Formulas
    new_rows = list(ws_source.iter_rows(min_row=2, max_col=25, values_only=True))
    if not new_rows:
        messagebox.showinfo("No Data", "No rows found in source file (starting at row 2).")
        return
    for row in new_rows:
        ws_dest.append(row)
    for i, _ in enumerate(new_rows, start=start_row):
        ws_dest[f"Z{i}"] = f"=IF(K{i}=0,\"\",M{i}/K{i})"
        ws_dest[f"AA{i}"] = f"=IF(Z{i}=\"\",\"\",ABS(Z{i}-$AB$2))"
    ws_dest["AB2"] = "=MEDIAN(Z:Z)"
    ws_dest["AC2"] = "=SUMIF(AA:AA,\">0\")"
    wb_dest.save(dest_file)
    messagebox.showinfo("Success", f"Rows appended to '{sheet_selection}' successfully!")

# Add New Sheets with Formulas and Formatting
def add_new_sheets():
    sheet1_name = simpledialog.askstring("Sheet Name", "Enter Name for Sheet 1:")
    if not sheet1_name:
        return
    sheet2_name = simpledialog.askstring("Sheet Name", "Enter Name for Sheet 2:")
    if not sheet2_name:
        return
    try:
        wb = load_workbook("Annual Sales.xlsx")
    except FileNotFoundError:
        messagebox.showerror("Error", "Annual Sales.xlsx not found in this folder.")
        return
    # Create Sheets
    ws1 = wb.create_sheet(title=sheet1_name)
    ws2 = wb.create_sheet(title=sheet2_name)
    # Formatting for Sheet 1
    bold = Font(bold=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    colorbold = Font(bold=True, color="425370")
    # Sheet 1 Content
    ws1["A1"] = "Assessment Roll"
    ws1["A3"] = "Valuation Date"
    ws1["A5"] = "Stony Point Sales"
    # Apply Borders and Font
    for cell in ["A1", "A3", "A5"]:
        ws1[cell].border = border
        ws1[cell].font = colorbold
    ws1["C2"] = "# of Completed Sales"
    ws1["D2"] = "Minimum Sales Price"
    ws1["E2"] = "Maximum Sales Price"
    ws1["F2"] = "Average Sale Price"
    ws1["G2"] = "Median Sales Price"
    ws1["C4"] = "Average AV/SP Ratio"
    ws1["D4"] = "Median AV/SP Ratio"
    ws1["E4"] = "Weighted Mean Ratio"
    ws1["F4"] = "C.O.D"
    ws1["G4"] = "Price Related Differential"
    ws1["C7"] = "Total # of Sales"
    ws1["D7"] = "Excluded # of Sales"
    # Apply Formatting and Formulas
    for col in ["C2", "D2", "E2", "F2", "G2", "C4", "D4", "E4", "F4", "G4", "C7", "D7"]:
        ws1[col].font = bold
        ws1[col].alignment = Alignment(horizontal="center")
        ws1[col].border = border
    for col in ["C3", "D3", "E3", "F3", "G3", "C5", "D5", "E5", "F5", "G5", "C8", "D8"]:
        ws1[col].border = border
    ws1["C3"] = f"=COUNT('{sheet2_name}'!Z:Z)"
    ws1["D3"] = f"=MIN('{sheet2_name}'!K:K)"
    ws1["E3"] = f"=MAX('{sheet2_name}'!K:K)"
    ws1["F3"] = f"=AVERAGE('{sheet2_name}'!K:K)"
    ws1["G3"] = f"=MEDIAN('{sheet2_name}'!K:K)"
    ws1["C5"] = f"=AVERAGEIF('{sheet2_name}'!Z:Z,\">0\")"
    ws1["D5"] = f"='{sheet2_name}'!AB2"
    ws1["E5"] = f"=SUM('{sheet2_name}'!M:M)/SUM('{sheet2_name}'!K:K)"
    ws1["F5"] = f"=((('{sheet2_name}'!AC2)/C3)/D5)*100"
    ws1["G5"] = "=C5/E5"
    ws1["C8"] = f"=COUNTA('{sheet2_name}'!B:B)-1"
    ws1["D8"] = f"=C8-C3"
    # Currency Formatting for Sheet 1
    ws1["D3"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws1["E3"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws1["F3"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws1["G3"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    # Sheet 2 Content
    headers = [
        "MLS #", "St", "Parcel Number", "Address", "Post Office/Town", "Lot Size Acres",
        "High School District", "City/Township", "Close Date", "Current Price",
        "5217 Sales Price", "Difference (Sales Price)", "5217 Assessed Value",
        "Current Assessed Value", "Difference (AV)", "Tax Assessed Value",
        "Year Built", "Living Sqft (Est)", "Bedrooms Total", "Bathrooms Full",
        "Bathrooms Half", "Cooling", "Association Fee", "Verified (Y/N)",
        "Condition Code", "AV/SP Ratio", "Absolute Deviation",
        "Median Ratio", "Sum Abs Dev"
    ]
    highlight_headers = [
        "5217 Sales Price", "5217 Assessed Value", "Current Assessed Value",
        "Verified (Y/N)", "Condition Code"
    ]
    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col)
        cell.value = header
        if header in highlight_headers:
            cell.font = bold
            cell.fill = highlight
    MAX_ROWS = 400
    # Formulas for Sheet 2
    for row in range(2, MAX_ROWS + 1):
        ws2[f"Z{row}"] = f"=IF(K{row}=0,\"\",M{row}/K{row})"
        ws2[f"AA{row}"] = f"=IF(Z{row}=\"\",\"\",ABS(Z{row}-$AB$2))"
    ws2["AB2"] = "=MEDIAN(Z:Z)"
    ws2["AC2"] = "=SUMIF(AA:AA,\">0\")"
    wb.save("Annual Sales.xlsx")
    messagebox.showinfo("Success", "Sheets added and formulas applied successfully!")

# Generate Bar Graph and Save as PNG
def generate_graph():
    try:
        df = pd.read_excel("Annual Sales.xlsx", sheet_name="Total Summary Analysis", usecols=["Assessment Roll", "Average Sales Price", "Median Sales Price"])
    except Exception as e:
        messagebox.showerror("Graph Error", f"Unable to read data.\n\n{e}")
        return
    df = df.set_index("Assessment Roll")
    fig, ax = plt.subplots(figsize=(11, 6))
    df.plot(kind="bar", ax=ax, width=0.3, color=["#e3782b", "#239130"])
    # Graph Formatting
    ax.set_title("Town of Stony Point Residential Average Sales Price vs. Median Sales Price\n(Single, Two - Four Family, and Condos Accounted For)", fontsize=10)
    ax.set_ylabel("Sales Price")
    ax.set_xlabel("Assessment Roll")
    ax.yaxis.set_major_formatter(StrMethodFormatter('${x:,.0f}'))
    for p in ax.patches:
        ax.text(p.get_x() + p.get_width() / 2, p.get_height(), f'${p.get_height():,.0f}', ha="center", va="bottom", fontsize=8)
    ax.tick_params(axis="y", labelleft=True)
    ax.tick_params(axis="x", labelbottom=True)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha="right")
    ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
    ax.grid(axis="y", linestyle="-", alpha=0.25)
    plt.tight_layout()
    filename = simpledialog.askstring("Save Graph", "Enter a filename (without extension):")
    if not filename:
        messagebox.showinfo("Cancelled", "Graph save cancelled.")
        return
    full_path = f"{filename}.png"
    try:
        plt.savefig(full_path, dpi=300)
        messagebox.showinfo("Success", f"Graph saved as:\n{full_path}")
    except Exception as e:
        messagebox.showerror("Save Error", f"Unable to save graph.\n\n{e}")

# Main GUI
def main():
    # Create Main Window
    root = tk.Tk()
    root.title("Property Data Processing Software")
    root.geometry("600x400")
    
    # Data Processing Frame
    frame1 = tk.LabelFrame(root, text="Data Processing", font=("Arial", 12, "bold"), padx=10, pady=10)
    frame1.pack(fill="x", padx=10, pady=10)
    # Data Processing Button
    btn1 = tk.Button(frame1, text="Process Data", command=run_processing, width=30, height=2)
    btn1.pack(pady=5)
    
    # Annual Sales Tools Frame
    frame2 = tk.LabelFrame(root, text="AnnualSales Tools", font=("Arial", 12, "bold"), padx=10, pady=10)
    frame2.pack(fill="x", padx=10, pady=10)
    # Annual Sales Tools Buttons
    # Append Rows Button
    btn2 = tk.Button(frame2, text="Append Rows", command=append_rows, width=30, height=2)
    btn2.pack(pady=5)
    # Add New Sheets Button
    btn3 = tk.Button(frame2, text="Add New Sheets", command=add_new_sheets, width=30, height=2)
    btn3.pack(pady=5)
    # Generate Chart Button
    btn4 = tk.Button(frame2, text="Generate Chart", command=generate_graph, width=30, height=2)
    btn4.pack(pady=5)

    root.mainloop()

# Run the Main Function
if __name__ == "__main__":
    main()
